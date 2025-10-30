import os, imaplib, email, io, json, re
from email.header import decode_header, make_header
from openpyxl import load_workbook
import gspread
from google.oauth2.service_account import Credentials

from datetime import datetime, date, timezone





SUBJECT_PHRASE = "Auto Glass Booked"
GMAIL_ADDRESS  = "apatilglassdoctordfw@gmail.com"
GMAIL_APP_PW   = "mvfgtidfegpgbbwo"
GSPREAD_SHEET_ID = "17e--bcU21i_XPxTJarE7evf2eDiQzWwbuKrR-nCVpk8"
GDRIVE_SA_JSON = r"C:\Users\Aayush Patil\Desktop\Daily Number Report\sheetsautomation-476714-3d5a94c2ed92.json"  # secret JSON string
# Google Sheet that holds the mapping (two columns: names, category)
GSPREAD_MAP_SHEET_ID = "1s60Pw4vDrQ4v97TQLBLwxlHmynOJSMm42OuSldN2W_w"
# optional: the tab name (default = first sheet)
MAP_TAB_NAME = "Sheet1"


def connect_imap():
    imap = imaplib.IMAP4_SSL("imap.gmail.com")
    imap.login(GMAIL_ADDRESS, GMAIL_APP_PW)
    imap.select("INBOX")
    return imap

def parse_money(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = re.sub(r'[^0-9.\-]', '', str(val))
    return float(s) if s else 0.0

def ensure_sheet1_header(ws):
    wanted = [[
        "Date","Auto Count","Auto Rev","Auto Warranties",
        "Flat Appointments","Flat with SD",
        "Callbacks/CPU/CSR/AceMan Count","Callbacks/CPU/CSR/AceMan Rev",
        "Sold jobs by specialists count","Sold jobs by specialists rev",
        "Total count","Total Rev","Leads","Booked"
    ]]
    hdr = ws.get_values("A1:N1")
    if not hdr or not hdr[0] or all((c or "").strip() == "" for c in hdr[0]):
        ws.update(wanted, range_name="A1", value_input_option="USER_ENTERED")



def upsert_sheet1(date_str, *, auto_count=None, auto_rev=None, auto_warr=None,
                  flat_appts=None, flat_with_sd=None,
                  cb_count=None, cb_rev=None, spec_count=None, spec_rev=None,
                  total_count=None, total_rev=None,
                  leads=None, booked=None):
    """
    Upsert one row in Sheet1 keyed by Date (col A).
    Only updates columns for which a value is provided.
    """
    gc = get_gspread_client()
    sh = gc.open_by_key(GSPREAD_SHEET_ID)
    try:
        ws = sh.worksheet("Sheet1")
    except gspread.WorksheetNotFound:
        ws = sh.get_worksheet(0)

    # Ensure A..N header exists
    ensure_sheet1_header(ws)

    # ---- find existing row first ----
    colA = ws.col_values(1)  # includes header
    target_row = None
    for i, v in enumerate(colA, start=1):
        if i == 1:
            continue
        if str(v).strip() == date_str:
            target_row = i
            break

    # Append new row if not found
    if target_row is None:
        row = [
            date_str,
            "" if auto_count   is None else auto_count,
            "" if auto_rev     is None else auto_rev,
            "" if auto_warr    is None else auto_warr,
            "" if flat_appts   is None else flat_appts,
            "" if flat_with_sd is None else flat_with_sd,
            "" if cb_count     is None else cb_count,
            "" if cb_rev       is None else cb_rev,
            "" if spec_count   is None else spec_count,
            "" if spec_rev     is None else spec_rev,
            "" if total_count  is None else total_count,
            "" if total_rev    is None else total_rev,
            "" if leads        is None else leads,
            "" if booked       is None else booked,
        ]
        ws.append_rows([stringify_matrix([row])[0]],
                       value_input_option="USER_ENTERED",
                       table_range="A1:N1")
        return

    # ---- update existing row (batch) ----
    updates = []

    def add(col_letter, value):
        if value is None:
            return
        updates.append({"range": f"{col_letter}{target_row}",
                        "values": [[to_cell(value)]]})

    add("A", date_str)
    add("B", auto_count)
    add("C", auto_rev)
    add("D", auto_warr)
    add("E", flat_appts)
    add("F", flat_with_sd)
    add("G", cb_count)
    add("H", cb_rev)
    add("I", spec_count)
    add("J", spec_rev)
    add("K", total_count)
    add("L", total_rev)
    add("M", leads)
    add("N", booked)

    if updates:
        ws.batch_update(updates, value_input_option="USER_ENTERED")



def compute_csr_metrics(rows, filename):
    """
    Returns: (date_str, leads, booked)
      - leads  = last non-empty in 'Lead Calls'
      - booked = last non-empty in 'Inbound Calls Booked'
    """
    dt = extract_date_from_filename(filename)
    if not rows or len(rows) < 2:
        return (dt, 0, 0)

    header = rows[0]
    body   = rows[1:]

    lead_col   = find_col_idx(header, {"lead calls", "leads"})
    booked_col = find_col_idx(header, {
        "inbound calls booked",
        "calls booked",
        "booked"
    })

    leads = 0
    if lead_col is not None:
        for r in reversed(body):
            v = r[lead_col]
            if v not in (None, "", " "):
                try:
                    leads = int(parse_money(v))
                except:
                    leads = int(parse_money(v))
                break

    booked = 0
    if booked_col is not None:
        for r in reversed(body):
            v = r[booked_col]
            if v not in (None, "", " "):
                try:
                    booked = int(parse_money(v))
                except:
                    booked = int(parse_money(v))
                break

    return (dt, leads, booked)



def load_name_category_map():
    """Returns dict: normalized_name -> category ('Callbacks/CPU/CSR/AccMan' or 'Sold jobs by specialists')"""
    gc = get_gspread_client()
    sh = gc.open_by_key(GSPREAD_MAP_SHEET_ID)
    ws = sh.worksheet(MAP_TAB_NAME) if MAP_TAB_NAME else sh.get_worksheet(0)
    values = ws.get_values("A1:B1000")  # [ ['names','category'], ... ]
    if not values or len(values) < 2:
        return {}
    header = [s.strip().lower() for s in values[0]]
    try:
        name_idx = header.index("names")
        cat_idx  = header.index("category")
    except ValueError:
        # assume first two cols
        name_idx, cat_idx = 0, 1
    m = {}
    for r in values[1:]:
        if len(r) <= max(name_idx, cat_idx):
            continue
        name = (r[name_idx] or "").strip().lower()
        cat  = (r[cat_idx] or "").strip()
        if name and cat:
            m[name] = cat
    return m

def compute_soldby_metrics(rows, filename, name_cat):
    """
    Returns tuple:
      (date_str, total_count, total_rev, cb_count, cb_rev, spec_count, spec_rev)
    """
    dt = extract_date_from_filename(filename)
    if not rows or len(rows) < 2:
        return (dt, 0, 0.0, 0, 0.0, 0, 0.0)

    header = rows[0]
    body   = rows[1:]

    soldby_col = find_col_idx(header, {"sold by", "soldby", "sold by name", "sold-by"})
    estid_col  = find_col_idx(header, {"estimate id", "estimateId".lower()})
    sub_col    = find_col_idx(header, {"estimates subtotal", "estimate subtotal", "subtotal"})

    # Totals = last non-empty values in those columns
    total_count = 0
    if estid_col is not None:
        for r in reversed(body):
            v = r[estid_col]
            if v not in (None, "", " "):
                # If it's a numeric running count, parse; otherwise coerce len
                try:
                    total_count = int(parse_money(v))
                except:
                    total_count = int(parse_money(v))
                break

    total_rev = 0.0
    if sub_col is not None:
        for r in reversed(body):
            v = r[sub_col]
            if v not in (None, "", " "):
                total_rev = parse_money(v)
                break

    # Category sums by mapping
    cb_count = cb_rev = spec_count = spec_rev = 0
    if soldby_col is not None and sub_col is not None:
        for r in body:
            name = (r[soldby_col] or "").strip().lower()
            if not name:
                continue
            cat = name_cat.get(name)
            if not cat:
                continue  # ignore unmapped names
            rev = parse_money(r[sub_col])
            if "callbacks" in cat.lower() or "cpu" in cat.lower() or "csr" in cat.lower() or "accman" in cat.lower():
                cb_count += 1
                cb_rev   += rev
            elif "specialist" in cat.lower():
                spec_count += 1
                spec_rev   += rev

    return (dt, total_count, total_rev, cb_count, cb_rev, spec_count, spec_rev)


def to_cell(v):
    # Convert non-JSON types to strings Google Sheets accepts
    if isinstance(v, (datetime, date)):
        # choose the format you want in the sheet:
        return v.strftime("%Y-%m-%d %H:%M:%S") if isinstance(v, datetime) else v.strftime("%Y-%m-%d")
    # (optional) handle Decimal etc. here if needed
    return v if v is not None else ""

def stringify_matrix(matrix):
    return [[to_cell(x) for x in row] for row in matrix]

def extract_date_from_filename(fname):
    # grabs the first pattern like 10_28_25 or 2025-10-28 etc.
    m = re.search(r'(\d{1,4})[._-](\d{1,2})[._-](\d{1,4})', fname)
    if not m:
        # fallback for “10_28_25 - 10_28_25”: use first group of 3 numbers
        m = re.search(r'(\d{1,2})[._-](\d{1,2})[._-](\d{2,4})', fname)
    if not m:
        return datetime.now(timezone.utc).date().isoformat()   # fallback = today UTC

    a,b,c = m.groups()
    # Normalize to YYYY-MM-DD assuming the short year is YY
    nums = list(map(int, (a,b,c)))
    if nums[0] > 31:  # likely YYYY-MM-DD
        yyyy, mm, dd = nums[0], nums[1], nums[2]
    else:             # likely MM-DD-YY
        mm, dd, yy = nums
        yyyy = 2000 + yy if yy < 100 else yy
    return f"{yyyy:04d}-{mm:02d}-{dd:02d}"

def find_col_idx(header, targets):
    """header: list[str]; targets: set|list of names (lowercase)"""
    h = [str(x).strip().lower() for x in header]
    for i, name in enumerate(h):
        if name in targets:
            return i
    return None

def compute_flat_metrics(rows, filename):
    """
    Returns: [date, flat_appointments, flat_with_sd]
      - date from filename
      - flat_appointments: last non-empty in 'Job #'
      - flat_with_sd: count of rows where 'Tags' column contains '$'
    """
    dt = extract_date_from_filename(filename)
    if not rows or len(rows) < 2:
        return [dt, 0, 0]

    header = rows[0]
    body   = rows[1:]

    job_col  = find_col_idx(header, {"job #", "job#", "jobs #"})
    tags_col = find_col_idx(header, {"tags"})

    # Last non-empty in Job #
    flat_appts = 0
    if job_col is not None:
        for r in reversed(body):
            if r and r[job_col] not in (None, "", " "):
                try:
                    flat_appts = int(parse_money(r[job_col]))
                except:
                    flat_appts = int(parse_money(r[job_col]))
                break

    # Count rows where Tags contains '$'
    flat_with_sd = 0
    if tags_col is not None:
        for r in body:
            cell = r[tags_col]
            if cell is not None and "$" in str(cell):
                flat_with_sd += 1

    return [dt, flat_appts, flat_with_sd]


def search_latest_matching(imap, phrase):
    """Find latest email whose subject contains `phrase`."""
    typ, data = imap.search(None, '(SUBJECT "{}")'.format(phrase))
    if typ != "OK":
        return None
    ids = data[0].split()
    return ids[-1] if ids else None


def get_first_xlsx_attachment(imap, msg_id):
    typ, data = imap.fetch(msg_id, "(RFC822)")
    if typ != "OK": 
        return None, None
    msg = email.message_from_bytes(data[0][1])
    subject = str(make_header(decode_header(msg.get("Subject", ""))))
    # walk parts
    for part in msg.walk():
        if part.get_content_disposition() == "attachment":
            fname = part.get_filename()
            if fname:
                fname = str(make_header(decode_header(fname)))
            if fname and fname.lower().endswith(".xlsx"):
                content = part.get_payload(decode=True)
                return fname, content
    return None, None

def read_xlsx_first_sheet(xlsx_bytes):
    wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = []
    for r in ws.iter_rows(values_only=True):
        rows.append([("" if v is None else v) for v in r])
    return rows  # 2D array

def compute_summary(rows, filename):
    """
    Returns a single row: [Date, Auto Count, Auto Rev, Auto Warranties]
    - Date from filename
    - Auto Count = last non-empty numeric in 'Job #' column
    - Auto Rev   = last non-empty numeric in 'Jobs Subtotal' column
    - Auto Warranties = count rows where any cell contains 'warranty'
    """
    if not rows or len(rows) < 2:
        dt = extract_date_from_filename(filename)
        return [[dt, 0, 0.0, 0]]

    header = rows[0]
    body   = rows[1:]

    job_col = find_col_idx(header, {"job #", "job#", "jobs #"})
    sub_col = find_col_idx(header, {"jobs subtotal", "job subtotal", "jobs sub total"})

    # last non-empty values in those columns
    job_count = 0
    if job_col is not None:
        for r in reversed(body):
            if r and r[job_col] not in (None, "", " "):
                try:
                    job_count = int(parse_money(r[job_col]))
                except:
                    job_count = int(parse_money(r[job_col]))
                break

    auto_rev = 0.0
    if sub_col is not None:
        for r in reversed(body):
            if r and r[sub_col] not in (None, "", " "):
                auto_rev = parse_money(r[sub_col])
                break

    # warranty count: any cell containing "warranty"
    warr_count = 0
    for r in body:
        row_text = " | ".join("" if c is None else str(c) for c in r).lower()
        if "warranty" in row_text:
            warr_count += 1

    dt = extract_date_from_filename(filename)
    return [[dt, job_count, auto_rev, warr_count]]


def get_gspread_client():
    raw = GDRIVE_SA_JSON  # can be a JSON string or a path to a .json file
    if os.path.exists(raw):
        with open(raw, "r", encoding="utf-8") as f:
            sa_info = json.load(f)
    else:
        sa_info = json.loads(raw)

    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(sa_info, scopes=scopes)
    return gspread.authorize(creds)

def append_to_sheet1(summary_row):
    """
    summary_row -> [[Date, Auto Count, Auto Rev, Auto Warranties]]
    Appends to Sheet1; creates the header once if A1:D1 is empty.
    """
    gc = get_gspread_client()
    sh = gc.open_by_key(GSPREAD_SHEET_ID)

    try:
        ws = sh.worksheet("Sheet1")
    except gspread.WorksheetNotFound:
        # Fallback: first sheet
        ws = sh.get_worksheet(0)

    # If A1:D1 is empty, write headers
    hdr = ws.get_values("A1:D1")
    if not hdr or not hdr[0] or all((c or "").strip() == "" for c in hdr[0]):
        ws.update([["Date", "Auto Count", "Auto Rev", "Auto Warranties"]],
                  range_name="A1",
                  value_input_option="USER_ENTERED")

    # Append the new row under the header
    ws.append_rows(
        stringify_matrix(summary_row),
        value_input_option="USER_ENTERED",
        table_range="A1:D1"   # keeps data aligned under headers
    )


def upsert_flat_into_sheet1(date_str, flat_appts, flat_with_sd):
    """If A column has date_str, update E and F on that row; else append a new row."""
    gc = get_gspread_client()
    sh = gc.open_by_key(GSPREAD_SHEET_ID)

    try:
        ws = sh.worksheet("Sheet1")
    except gspread.WorksheetNotFound:
        ws = sh.get_worksheet(0)

    # Ensure header exists
    hdr = ws.get_values("A1:N1")
    if not hdr or not hdr[0] or all((c or "").strip() == "" for c in hdr[0]):
        ws.update([["Date","Auto Count","Auto Rev","Auto Warranties",
                    "Flat Appointments","Flat with SD","Callbacks/CPU/CSR/AceMan Count",
                    "Callbacks/CPU/CSR/AceMan Rev","Sold jobs by specialists count",
                    "Sold jobs by specialists rev","Total count","Total Rev","Leads","Booked"]],
                  range_name="A1", value_input_option="USER_ENTERED")

    # Find existing row by exact match in Col A
    colA = ws.col_values(1)  # includes header
    target_row = None
    for idx, val in enumerate(colA, start=1):
        if idx == 1:
            continue
        if str(val).strip() == date_str:
            target_row = idx
            break

    if target_row:
        # Update columns E and F in that row
        ws.update([[flat_appts, flat_with_sd]],
                  range_name=f"E{target_row}:F{target_row}",
                  value_input_option="USER_ENTERED")
    else:
        # Append a new line: A..F with blanks for B..D
        ws.append_rows([ [date_str, "", "", "", flat_appts, flat_with_sd] ],
                       value_input_option="USER_ENTERED",
                       table_range="A1:F1")



def main():
    imap = connect_imap()
    try:
        msg_id = search_latest_matching(imap, "Auto Glass Booked")
        if not msg_id:
            print("No matching emails found.")
            return
        fname, content = get_first_xlsx_attachment(imap, msg_id)
        if not content:
            print("Email found, but no .xlsx attachment.")
            return
        rows = read_xlsx_first_sheet(content)
        summary = compute_summary(rows, fname)   # [[date, count, rev, warranties]]
        date_str, count, rev, warr = summary[0]
        upsert_sheet1(date_str, auto_count=count, auto_rev=rev, auto_warr=warr)

        print(f"Processed {fname} | rows={len(rows)}")
    # --- New: Flat Opportunities Booked block ---
        msg_id_flat = search_latest_matching(imap, "Flat Opportunities Booked")
        if msg_id_flat:
            fname2, content2 = get_first_xlsx_attachment(imap, msg_id_flat)
            if content2:
                rows2 = read_xlsx_first_sheet(content2)
                date_str, flat_appts, flat_with_sd = compute_flat_metrics(rows2, fname2)
                upsert_sheet1(date_str, flat_appts=flat_appts, flat_with_sd=flat_with_sd)
                print(f"Upserted Flat metrics for {date_str}: appts={flat_appts}, withSD={flat_with_sd}")
        # --- Sold by Report ---
        msg_id_sold = search_latest_matching(imap, "Sold by Report")
        if msg_id_sold:
            fname3, content3 = get_first_xlsx_attachment(imap, msg_id_sold)
            if content3:
                rows3 = read_xlsx_first_sheet(content3)
                name_cat = load_name_category_map()  # read mapping sheet once
                (date_str, total_count, total_rev,
                 cb_count, cb_rev, spec_count, spec_rev) = compute_soldby_metrics(rows3, fname3, name_cat)

                upsert_sheet1(date_str,
                              cb_count=cb_count, cb_rev=cb_rev,
                              spec_count=spec_count, spec_rev=spec_rev,
                              total_count=total_count, total_rev=total_rev)
                print(f"Upserted Sold-by metrics for {date_str}: total_count={total_count}, total_rev={total_rev}")
        # --- CSR Bookings Report ---
        msg_id_csr = search_latest_matching(imap, "CSR Bookings Report")
        if msg_id_csr:
            fname4, content4 = get_first_xlsx_attachment(imap, msg_id_csr)
            if content4:
                rows4 = read_xlsx_first_sheet(content4)
                date_str, leads, booked = compute_csr_metrics(rows4, fname4)
                upsert_sheet1(date_str, leads=leads, booked=booked)
                print(f"Upserted CSR metrics for {date_str}: leads={leads}, booked={booked}")

    finally:
        try:
            imap.close()
        except Exception:
            pass
        imap.logout()

if __name__ == "__main__":
    main()
